import re
import datetime
from openpyxl import Workbook
from selenium import webdriver

upid_map = {
    '北京市': '1', '天津市': '2', '河北省': '3', '山西省': '4', '内蒙古': '5',
    '辽宁省': '6', '吉林省': '7', '黑龙江': '8', '上海市': '9', '江苏省': '10',
    '浙江省': '11', '安徽省': '12', '福建省': '13', '江西省': '14', '山东省': '15',
    '河南省': '16', '湖北省': '17', '湖南省': '18', '广东省': '19', '广西': '20',
    '海南省': '21', '重庆市': '22', '四川省': '23', '贵州省': '24', '云南省': '25',
    '西藏': '26', '陕西省': '27', '甘肃省': '28', '青海省': '29', '宁夏': '30',
    '新疆': '31', '台湾省': '32', '香港': '33', '澳门': '34', '海外': '35',
    '其他': '36'}

data_xls_poz_map = {
    1: {'name': '页面链接', 'code': 'url'},
    2: {'name': '标题', 'code': 'title'},
    3: {'name': '发布日期', 'code': 'public_date'},
    4: {'name': '信息类型', 'code': 'info_type'},
    5: {'name': '区域名称', 'code': 'region_name'},
    6: {'name': '类别从属', 'code': 'type_depc'},
    7: {'name': '付款方式', 'code': 'pay_way'},
    8: {'name': '价格', 'code': 'price'},
    9: {'name': '数量', 'code': 'acount'},
    10: {'name': '出肉率', 'code': 'meat_percent'},
    11: {'name': '体重', 'code': 'weight'},
    12: {'name': '毛色', 'code': 'skin_color'},
    13: {'name': '发布人', 'code': 'publisher'},
    14: {'name': '联系人', 'code': 'contact'},
    15: {'name': '公司名字', 'code': 'company'},
    16: {'name': '联系电话', 'code': 'telphone'},
    17: {'name': 'QQ咨询', 'code': 'qq_number'},
    18: {'name': '猪场类型', 'code': 'farm_type'},
    19: {'name': '发布者IP', 'code': 'pub_ip'}
}

base_url = 'http://bbs.zhue.com.cn/gongqiu.php'  # Target website base url
query_mode = 'index'  # query type
query_level = '1'  # data level
query_upid = upid_map['吉林省']  # Province Jilin
today = datetime.date.today()  # start date
now = datetime.datetime.now()  # start datetime
file_name = '生猪供求数据' + str(today) + '.xlsx'  # data-container file
current_page = 1
page_info_list = []  # data list

brower = webdriver.Firefox()  # use FireFox Brower


# url constructor
def url_constructor(page_num):
    return base_url + '?mod=' + query_mode + '&level=' + query_level \
           + '&upid=' + query_upid + '&page=' + str(page_num)


# get url-list
def details_url_list_getter(url):
    brower.get(url)
    link_list = []
    span_list = brower.find_elements_by_xpath('//*[@class="gongqiu_pic tc"]')
    for span in span_list:
        link_list.append(span.find_element_by_tag_name('a').get_attribute('href'))
    return link_list


# get details
def details_info_getter(details_url):
    brower.get(details_url)
    # include `TITLE` and `PUBLIC DATE`
    top_block = brower.find_element_by_xpath('//*[@class="index_content_title"]')
    # title
    title = top_block.find_element_by_tag_name('h1').text
    # publish_date
    publish_date = top_block.find_element_by_xpath('//*[@class="index_content_title_subtitle"]').text
    publish_date = re.findall(r"(\d{2}-\d{1,2}-\d{1,2}\s\d{1,2}:\d{1,2})", publish_date)[0]
    # info table
    content_block = brower.find_element_by_xpath('//*[@class="index_content_extracontact"]')
    # info spans
    info_spans = content_block.find_elements_by_xpath('//*[@class="val"]')
    # ip location
    ip_location = content_block.find_element_by_xpath('//*[@class="ip"]')
    try:
        result_dic = {}
        result_dic['url'] = details_url  # 页面链接
        result_dic['title'] = title  # 标题
        result_dic['public_date'] = publish_date  # 发布日期
        result_dic['info_type'] = info_spans[0].text  # 信息类型
        result_dic['region_name'] = info_spans[1].text  # 区域名称
        result_dic['type_depc'] = info_spans[2].text  # 类别从属
        result_dic['pay_way'] = info_spans[3].text  # 付款方式
        result_dic['price'] = info_spans[4].text  # 价格
        result_dic['acount'] = info_spans[5].text  # 数量
        result_dic['meat_percent'] = info_spans[6].text  # 出肉率
        result_dic['weight'] = info_spans[7].text  # 体重
        result_dic['skin_color'] = info_spans[8].text  # 毛色
        result_dic['publisher'] = info_spans[9].text  # 发布人
        result_dic['contact'] = info_spans[10].text  # 联系人
        result_dic['company'] = info_spans[11].text  # 公司名字
        result_dic['telphone'] = info_spans[12].text  # 联系电话
        result_dic['qq_number'] = info_spans[13].text  # QQ咨询
        result_dic['farm_type'] = info_spans[14].text  # 猪场类型
        result_dic['pub_ip'] = info_spans[15].text + ip_location.text  # 发布者IP
    except IndexError:
        print("--------链接-------")
        print(details_url)
        print("--------Span-------")
        print(info_spans)
        print("--------Block-------")
        print(content_block)
        print("\n\n")
    return result_dic


# data output to `xls` file
def data_output_xls(data_list, current_page):
    print('开始写文件(第'+str(current_page)+'页)....')
    wb = Workbook()
    sheet_names = wb.get_sheet_names()
    work_sheet = wb.get_sheet_by_name(sheet_names[0])
    for column in range(1, len(data_xls_poz_map)):
        _ = work_sheet.cell(column=column, row=1, value="%s" % data_xls_poz_map[column]['name'])
    for row in range(2, len(data_list)):
        for column in range(1, len(data_xls_poz_map)):
            _ = work_sheet.cell(column=column, row=row, value="%s" % data_list[row][data_xls_poz_map[column]['code']])
    # save begin
    try:
        wb.save(filename=file_name)
        print('完成写文件(第'+str(current_page)+'页)....')
    except IOError as iox:
        print('文件读写异常')
        print(data_list)  # 将数据输出,避免数据因异常丢失
        print('错误信息:')
        for e in iox.args:
            print(e)
        print('数据写入文件失败....')
    except Exception as unknown:
        print("未知异常导致文件输出失败!错误信息:")
        for e in unknown.args:
            print(e)


# spider logic...
def spider(current_page=1):
    while True:
        url_list = details_url_list_getter(url_constructor(current_page))
        if len(url_list) > 0:
            print('开始获取第' + str(current_page) + '页数据...')
            for url in url_list:
                page_info_list.append(details_info_getter(url))
            print('完成获取第' + str(current_page) + '页数据...')
            data_output_xls(page_info_list, current_page)
            current_page += 1
        else:
            break


# Main method
if __name__ == '__main__':
    try:
        spider(current_page)
        print("完成本次数据爬取任务!")
    finally:
        brower.close()
